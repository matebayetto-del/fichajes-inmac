
# -*- coding: utf-8 -*-
"""
Planilla universal de fichajes / horas

Qué hace:
- Detecta automáticamente cuál archivo es la planilla base y cuál es el archivo de fichajes.
- No depende de nombres fijos de archivo.
- Detecta el/los mes/es presentes en fichajes y genera una salida por cada mes.
- Reconfigura los encabezados de la planilla base al mes detectado.
- Reescribe las fórmulas de:
    * horas primera quincena
    * horas segunda quincena
    * horas totales
  según el calendario real del mes, contemplando:
    * horas normales: columnas verdes lunes a sábado
    * horas al 50%: columnas amarillas lunes a viernes
    * horas al 100%:
        - columna amarilla de sábados
        - columnas verdes y amarillas de domingos y feriados
- Mantiene o reescribe las fórmulas diarias según el tipo de plantilla detectada.
- Carga entrada/salida como valores de hora reales de Excel (no como texto).
- Genera hojas auxiliares de revisión.

Uso en Colab:
1) Ejecutar la celda.
2) Subir los 2 Excel.
3) El script detecta qué archivo es cada uno y genera la/s planilla/s de salida.
"""

# =========================================================
# INSTALACIÓN DE PAQUETES
# =========================================================
import sys
import subprocess
import importlib.util

def _instalar_si_falta(paquete, import_name=None):
    import_name = import_name or paquete
    if importlib.util.find_spec(import_name) is None:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", paquete])

for pkg_name, import_name in [
    ("pandas", "pandas"),
    ("openpyxl", "openpyxl"),
    ("xlrd", "xlrd"),
    ("holidays", "holidays"),
]:
    _instalar_si_falta(pkg_name, import_name)

# =========================================================
# IMPORTS
# =========================================================
import os
import re
import copy
import math
import calendar
import unicodedata
from datetime import datetime, date, time
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
import holidays
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

try:
    from google.colab import files
except Exception:
    files = None


# =========================================================
# CONFIGURACIÓN
# =========================================================
UMBRAL_MATCH = 0.76
UMBRAL_DUDOSO = 0.88
UMBRAL_MINIMO_SIM_POR_LEGAJO = 0.50
TOLERANCIA_MINUTOS = 5
PROCESAR_TODOS_LOS_MESES = True

# Argentina por defecto. Cambiar si necesitás otra subdivisión.
PAIS_FERIADOS = "AR"
SUBDIV_FERIADOS = None

# Para sumar o excluir feriados manualmente:
# formato ISO: "YYYY-MM-DD"
FERIADOS_ADICIONALES = set()
FERIADOS_EXCLUIDOS = set()

# Qué días se marcan AUSENTE si no hay fichaje
# 0=lunes ... 6=domingo
DIAS_LABORABLES_AUSENCIA = {0, 1, 2, 3, 4, 5}  # lunes a sábado

# Equivalencias manuales opcionales
# clave = nombre en planilla
# valor = nombre como figura en fichajes
EQUIVALENCIAS_MANUALES = {
    # "ARCE RODRIGUEZ CELSO CARLOS": "celso arce",
}

# Nombre base de salida
PREFIJO_SALIDA = "PLANILLA_COMPLETA"

# Colores
FILL_DUDOSO = PatternFill(fill_type="solid", fgColor="FFF2CC")
FILL_NO_MATCH = PatternFill(fill_type="solid", fgColor="F4CCCC")
FILL_CARGADO = PatternFill(fill_type="solid", fgColor="D9EAD3")
FILL_INCIDENCIA = PatternFill(fill_type="solid", fgColor="FCE5CD")
NO_FILL = PatternFill(fill_type=None)


# =========================================================
# TEXTO / NOMBRES
# =========================================================
def quitar_tildes(texto):
    return "".join(
        c for c in unicodedata.normalize("NFD", str(texto))
        if unicodedata.category(c) != "Mn"
    )

def limpiar_texto_base(texto):
    if texto is None or (isinstance(texto, float) and math.isnan(texto)):
        return ""
    texto = str(texto).strip()
    texto = quitar_tildes(texto).lower()
    texto = re.sub(r"[,;:/\-\.\(\)\[\]{}]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto

def normalizar_texto(texto):
    texto = limpiar_texto_base(texto)
    ignorar = {"sr", "sra", "dr", "dra", "lic", "ing"}
    tokens = [t for t in texto.split() if t not in ignorar]
    tokens.sort()
    return " ".join(tokens)

def extraer_tokens_nombre(texto):
    return [t for t in limpiar_texto_base(texto).split() if t]

def apellido_y_primer_nombre(texto):
    tokens = extraer_tokens_nombre(texto)
    if not tokens:
        return ""
    if len(tokens) == 1:
        return tokens[0]
    firma = f"{tokens[-1]} {tokens[0]}"
    return " ".join(sorted(firma.split()))

def tokens_set(texto):
    return set(normalizar_texto(texto).split())

def normalizar_header(texto):
    return limpiar_texto_base(texto).replace(".", "").replace("%", "")

def normalizar_legajo(valor):
    if valor is None:
        return None
    if isinstance(valor, float) and math.isnan(valor):
        return None

    texto = str(valor).strip()
    if not texto:
        return None

    digitos = re.sub(r"\D", "", texto)
    if not digitos:
        return None

    return str(int(digitos))

def derivar_legajo_desde_id(valor_id):
    if valor_id is None:
        return None
    if isinstance(valor_id, float) and math.isnan(valor_id):
        return None

    digitos = re.sub(r"\D", "", str(valor_id).strip())
    if not digitos:
        return None

    legajo = digitos[3:] if len(digitos) > 3 else digitos
    legajo = legajo.lstrip("0") or "0"
    return legajo

def primer_valor_no_vacio(valores):
    for valor in valores:
        if valor is None:
            continue
        if isinstance(valor, float) and math.isnan(valor):
            continue
        if str(valor).strip() == "":
            continue
        return valor
    return None


# =========================================================
# MATCH DE NOMBRES
# =========================================================
def similitud_base(a, b):
    return SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio()

def similitud_token_sort(a, b):
    na = normalizar_texto(a)
    nb = normalizar_texto(b)
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()

def similitud_tokens(a, b):
    ta = tokens_set(a)
    tb = tokens_set(b)
    if not ta or not tb:
        return 0.0
    inter = len(ta & tb)
    union = len(ta | tb)
    return inter / union if union else 0.0

def similitud_apellido_primer_nombre(a, b):
    fa = apellido_y_primer_nombre(a)
    fb = apellido_y_primer_nombre(b)
    if not fa or not fb:
        return 0.0
    return SequenceMatcher(None, fa, fb).ratio()

def mismo_apellido(a, b):
    ta = extraer_tokens_nombre(a)
    tb = extraer_tokens_nombre(b)
    if not ta or not tb:
        return False
    return ta[-1] == tb[-1] or ta[0] == tb[0] or ta[-1] == tb[0] or ta[0] == tb[-1]

def mejor_match(nombre_objetivo, candidatos):
    mejor_nombre = None
    mejor_score = 0.0
    mejor_detalle = {}

    for candidato in candidatos:
        s1 = similitud_base(nombre_objetivo, candidato)
        s2 = similitud_token_sort(nombre_objetivo, candidato)
        s3 = similitud_tokens(nombre_objetivo, candidato)
        s4 = similitud_apellido_primer_nombre(nombre_objetivo, candidato)
        bonus_apellido = 0.03 if mismo_apellido(nombre_objetivo, candidato) else 0.0

        score = min(
            1.0,
            max(
                s1 * 0.85,
                s2,
                s3 * 0.95,
                s4 + bonus_apellido
            )
        )

        if score > mejor_score:
            mejor_score = score
            mejor_nombre = candidato
            mejor_detalle = {
                "base": round(s1, 4),
                "token_sort": round(s2, 4),
                "tokens": round(s3, 4),
                "apellido_primernombre": round(s4, 4),
                "bonus_apellido": round(bonus_apellido, 4),
            }

    return mejor_nombre, mejor_score, mejor_detalle

def buscar_equivalencia_manual(nombre_planilla, candidatos_fichajes):
    if not nombre_planilla:
        return None

    nombre_normalizado = normalizar_texto(nombre_planilla)

    equivalencias_norm = {
        normalizar_texto(k): normalizar_texto(v)
        for k, v in EQUIVALENCIAS_MANUALES.items()
    }

    if nombre_normalizado not in equivalencias_norm:
        return None

    objetivo_manual = equivalencias_norm[nombre_normalizado]

    for cand in candidatos_fichajes:
        if normalizar_texto(cand) == objetivo_manual:
            return cand

    return None


def resolver_coincidencia_persona(nombre_planilla, legajo_planilla, candidatos_fichajes, legajo_a_candidatos):
    nombre_planilla_limpio = normalizar_texto(nombre_planilla)
    legajo_planilla_norm = normalizar_legajo(legajo_planilla)

    match_manual = buscar_equivalencia_manual(nombre_planilla, candidatos_fichajes)
    if match_manual is not None:
        return {
            "nombre_match": match_manual,
            "score": 1.0,
            "score_original": 1.0,
            "detalle": {
                "base": 1.0,
                "token_sort": 1.0,
                "tokens": 1.0,
                "apellido_primernombre": 1.0,
                "manual": True,
                "metodo": "manual",
                "legajo_planilla": legajo_planilla_norm,
                "legajo_fichajes": legajo_planilla_norm,
                "candidatos_legajo": match_manual,
            },
            "aceptado": True,
            "dudoso": False,
        }

    nombre_match, score_nombre, detalle_nombre = mejor_match(nombre_planilla_limpio, candidatos_fichajes)
    detalle_nombre = dict(detalle_nombre or {})
    detalle_nombre.update({
        "manual": False,
        "metodo": "nombre",
        "legajo_planilla": legajo_planilla_norm,
        "legajo_fichajes": None,
        "candidatos_legajo": "",
    })

    candidatos_legajo = []
    if legajo_planilla_norm:
        candidatos_legajo = sorted(legajo_a_candidatos.get(legajo_planilla_norm, []))

    if candidatos_legajo:
        nombre_match_leg, score_leg, detalle_leg = mejor_match(nombre_planilla_limpio, candidatos_legajo)
        detalle_leg = dict(detalle_leg or {})
        detalle_leg.update({
            "manual": False,
            "legajo_planilla": legajo_planilla_norm,
            "legajo_fichajes": legajo_planilla_norm,
            "candidatos_legajo": ", ".join(candidatos_legajo),
            "cantidad_candidatos_legajo": len(candidatos_legajo),
        })

        mismo_candidato = nombre_match_leg and (nombre_match_leg == nombre_match)

        if mismo_candidato and score_nombre >= UMBRAL_MATCH:
            detalle_leg["metodo"] = "nombre+legajo"
            return {
                "nombre_match": nombre_match_leg,
                "score": score_nombre,
                "score_original": score_nombre,
                "detalle": detalle_leg,
                "aceptado": True,
                "dudoso": score_nombre < UMBRAL_DUDOSO,
            }

        if nombre_match_leg and score_leg >= UMBRAL_MINIMO_SIM_POR_LEGAJO:
            detalle_leg["metodo"] = "legajo" if len(candidatos_legajo) == 1 else "legajo+nombre"
            return {
                "nombre_match": nombre_match_leg,
                "score": max(score_leg, UMBRAL_MATCH),
                "score_original": score_leg,
                "detalle": detalle_leg,
                "aceptado": True,
                "dudoso": (score_leg < UMBRAL_DUDOSO) or (len(candidatos_legajo) > 1),
            }

    return {
        "nombre_match": nombre_match,
        "score": score_nombre,
        "score_original": score_nombre,
        "detalle": detalle_nombre,
        "aceptado": bool(nombre_match and score_nombre >= UMBRAL_MATCH),
        "dudoso": bool(nombre_match and (UMBRAL_MATCH <= score_nombre < UMBRAL_DUDOSO)),
    }


# =========================================================
# DETECCIÓN DE ARCHIVOS
# =========================================================
def _leer_preview_tabular(path, nrows=5):
    try:
        return pd.read_excel(path, nrows=nrows)
    except Exception:
        return None

def score_como_fichajes(path):
    score = 0
    df = _leer_preview_tabular(path, nrows=5)
    if df is None:
        return score

    cols = [normalizar_header(c) for c in df.columns]
    if any("nombre" in c for c in cols):
        score += 2
    if any(("hora" in c or "fecha" in c) for c in cols):
        score += 2
    if any("inout" in c for c in cols):
        score += 1
    if any("modo verificar" in c for c in cols):
        score += 1
    if len(cols) > 30:
        score -= 2
    return score

def score_como_planilla(path):
    ext = Path(path).suffix.lower()
    if ext not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return 0

    score = 0
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        for ws in wb.worksheets[:3]:
            max_r = min(5, ws.max_row)
            max_c = min(250, ws.max_column)
            textos = []
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    textos.append(str(ws.cell(r, c).value or ""))
            flat = normalizar_texto(" ".join(textos))
            if "entrada" in flat and "salida" in flat and "ausente" in flat and "horas" in flat:
                score += 3
            if "apellido y nombre" in flat or "nombre y apellido" in flat:
                score += 2
            if "horas primera quincena" in flat or "horas segunda quincena" in flat:
                score += 2
        wb.close()
    except Exception:
        pass
    return score

def detectar_archivos(archivos):
    if len(archivos) < 2:
        raise ValueError("Tenés que subir al menos 2 archivos Excel.")

    candidatos = []
    for path in archivos:
        if not Path(path).suffix.lower() in {".xls", ".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        candidatos.append({
            "path": path,
            "score_fichajes": score_como_fichajes(path),
            "score_planilla": score_como_planilla(path),
        })

    if len(candidatos) < 2:
        raise ValueError("No encontré suficientes archivos Excel válidos.")

    archivo_fichajes = None
    archivo_planilla = None

    mejor_fichajes = max(candidatos, key=lambda x: x["score_fichajes"])
    mejor_planilla = max(candidatos, key=lambda x: x["score_planilla"])

    if mejor_fichajes["score_fichajes"] <= 0:
        raise ValueError("No pude identificar automáticamente el archivo de fichajes.")
    if mejor_planilla["score_planilla"] <= 0:
        raise ValueError("No pude identificar automáticamente la planilla base.")

    if mejor_fichajes["path"] == mejor_planilla["path"]:
        raise ValueError(
            "El mismo archivo parece cumplir ambos roles. Revisá la estructura o subí solo los 2 Excel correctos."
        )

    archivo_fichajes = mejor_fichajes["path"]
    archivo_planilla = mejor_planilla["path"]

    return archivo_planilla, archivo_fichajes, candidatos


# =========================================================
# LECTURA DE FICHAJES
# =========================================================
def leer_fichajes(path_excel):
    df = pd.read_excel(path_excel)

    col_nombre = None
    col_hora = None
    col_inout = None
    col_id = None

    for c in df.columns:
        cl = normalizar_header(c)
        if col_nombre is None and "nom" in cl:
            col_nombre = c
        if col_hora is None and ("hora" in cl or "fecha" in cl):
            col_hora = c
        if col_inout is None and "inout" in cl:
            col_inout = c
        if col_id is None and (cl == "id" or "legajo" in cl):
            col_id = c

    if col_nombre is None:
        raise ValueError("No se encontró columna de nombre en fichajes.")
    if col_hora is None:
        raise ValueError("No se encontró columna de fecha/hora en fichajes.")

    rename_map = {col_nombre: "Nombre", col_hora: "Hora"}
    if col_inout is not None:
        rename_map[col_inout] = "Inout"
    if col_id is not None:
        rename_map[col_id] = "ID"

    df = df.rename(columns=rename_map)

    df["Hora"] = pd.to_datetime(df["Hora"], errors="coerce")
    df = df.dropna(subset=["Hora"]).copy()

    if df.empty:
        raise ValueError("No hay fichajes válidos.")

    if "ID" in df.columns:
        df["LegajoFichaje"] = df["ID"].apply(derivar_legajo_desde_id)
    else:
        df["LegajoFichaje"] = None

    df["Fecha"] = df["Hora"].dt.date
    df["Periodo"] = df["Hora"].dt.to_period("M")
    df["NombreOriginal"] = df["Nombre"].astype(str).str.strip()
    df["NombreLimpio"] = df["NombreOriginal"].apply(normalizar_texto)

    agrupado = (
        df.groupby(["Periodo", "NombreLimpio", "Fecha"], as_index=False)
        .agg(
            Entrada=("Hora", "min"),
            Salida=("Hora", "max"),
            CantidadFichajes=("Hora", "count"),
            NombreFichaje=("NombreOriginal", "first"),
            LegajoFichaje=("LegajoFichaje", lambda s: primer_valor_no_vacio(s)),
        )
    )

    return agrupado, df


# =========================================================
# DETECCIÓN DE ESTRUCTURA DE PLANILLA
# =========================================================

def detectar_fila_subtitulos(ws):
    mejor_fila = None
    mejor_score = -1

    for r in range(1, min(ws.max_row, 8) + 1):
        score = 0
        for c in range(1, min(ws.max_column, 250) + 1):
            valor = normalizar_header(ws.cell(r, c).value)
            if valor == "entrada":
                score += 3
            elif valor == "salida":
                score += 3
            elif valor == "ausente":
                score += 2
            elif valor == "horas":
                score += 1
            elif valor in {"horas 50", "horas50", "horas al 50", "horas al50"}:
                score += 1
            elif valor in {"horas 100", "horas100", "horas al 100", "horas al100"}:
                score += 1
            elif "horas ex" in valor or valor == "horas ex":
                score += 1
        if score > mejor_score:
            mejor_score = score
            mejor_fila = r

    if mejor_fila is None or mejor_score < 10:
        raise ValueError(
            "No pude detectar la fila de subtítulos (ENTRADA/SALIDA/AUSENTE/HORAS/HORAS 50%/HORAS 100% o HORAS EX.)."
        )

    return mejor_fila

def detectar_columna_nombre(ws):
    mejores = []
    for fila_header in range(1, min(ws.max_row, 6) + 1):
        for c in range(1, min(ws.max_column, 50) + 1):
            v = normalizar_header(ws.cell(fila_header, c).value)
            score = 0
            if "apellido y nombre" in v:
                score += 5
            if "nombre y apellido" in v:
                score += 5
            if "apellido" in v:
                score += 2
            if "nombre" in v:
                score += 2
            if score:
                mejores.append((score, fila_header, c))

    if mejores:
        mejores.sort(reverse=True)
        _, fila, col = mejores[0]
        return fila, col

    # fallback: fila 1, columna 5
    return 1, 5

def detectar_columna_legajo(ws):
    mejores = []
    for fila_header in range(1, min(ws.max_row, 6) + 1):
        for c in range(1, min(ws.max_column, 30) + 1):
            v = normalizar_header(ws.cell(fila_header, c).value)
            score = 0
            if v == "leg":
                score += 6
            if v == "legajo":
                score += 6
            if "legajo" in v:
                score += 4
            if re.fullmatch(r"leg", v):
                score += 4
            if score:
                mejores.append((score, fila_header, c))

    if mejores:
        mejores.sort(reverse=True)
        _, fila, col = mejores[0]
        return fila, col

    return None, None


def detectar_bloques_dia(ws, fila_fechas, fila_subtitulos):
    bloques = []

    c = 2
    while c <= ws.max_column - 4:
        v0 = normalizar_header(ws.cell(fila_subtitulos, c).value)
        v1 = normalizar_header(ws.cell(fila_subtitulos, c + 1).value)
        v2 = normalizar_header(ws.cell(fila_subtitulos, c + 2).value)
        v3 = normalizar_header(ws.cell(fila_subtitulos, c + 3).value)
        v4 = normalizar_header(ws.cell(fila_subtitulos, c + 4).value)
        v5 = normalizar_header(ws.cell(fila_subtitulos, c + 5).value) if c + 5 <= ws.max_column else ""

        if (
            v0 == "entrada"
            and v1 == "salida"
            and v2 == "ausente"
            and v3 == "horas"
            and ("horas ex" in v4 or v4 == "horas ex")
        ):
            base_col = c - 1
            bloques.append({
                "indice": len(bloques) + 1,
                "tipo_bloque": "5col",
                "ancho_bloque": 6,
                "base_col": base_col,
                "entrada": c,
                "salida": c + 1,
                "ausente": c + 2,
                "horas": c + 3,
                "horas_ex": c + 4,
                "header": ws.cell(fila_fechas, base_col).value,
            })
            c += 5
            continue

        if (
            v0 == "entrada"
            and v1 == "salida"
            and v2 == "ausente"
            and v3 == "horas"
            and v4 in {"horas 50", "horas50", "horas al 50", "horas al50"}
            and v5 in {"horas 100", "horas100", "horas al 100", "horas al100"}
        ):
            base_col = c - 1
            bloques.append({
                "indice": len(bloques) + 1,
                "tipo_bloque": "6col",
                "ancho_bloque": 7,
                "base_col": base_col,
                "entrada": c,
                "salida": c + 1,
                "ausente": c + 2,
                "horas": c + 3,
                "horas_50": c + 4,
                "horas_100": c + 5,
                "header": ws.cell(fila_fechas, base_col).value,
            })
            c += 6
            continue

        c += 1

    if not bloques:
        raise ValueError("No pude detectar los bloques diarios.")

    return bloques

def detectar_columnas_resumen(ws, fila_fechas):
    grupos = {}
    for c in range(1, ws.max_column + 1):
        valor = normalizar_header(ws.cell(fila_fechas, c).value)
        if not valor:
            continue
        if "horas totales" in valor:
            grupos["totales"] = c
        elif "primera quincena" in valor:
            grupos["q1"] = c
        elif "segunda quincena" in valor:
            grupos["q2"] = c

    if not all(k in grupos for k in ("totales", "q1", "q2")):
        raise ValueError("No pude detectar los grupos de resumen (totales / quincenas).")

    return {
        "totales": {"normales": grupos["totales"], "50": grupos["totales"] + 1, "100": grupos["totales"] + 2},
        "q1": {"normales": grupos["q1"], "50": grupos["q1"] + 1, "100": grupos["q1"] + 2},
        "q2": {"normales": grupos["q2"], "50": grupos["q2"] + 1, "100": grupos["q2"] + 2},
    }


def detectar_estructura_planilla(ws):
    fila_subtitulos = detectar_fila_subtitulos(ws)
    fila_fechas = max(1, fila_subtitulos - 1)

    fila_header, col_nombre = detectar_columna_nombre(ws)
    _, col_legajo = detectar_columna_legajo(ws)
    fila_inicio_datos = fila_subtitulos + 1

    bloques_dia = detectar_bloques_dia(ws, fila_fechas, fila_subtitulos)
    columnas_resumen = detectar_columnas_resumen(ws, fila_fechas)

    tipo_bloque = bloques_dia[0].get("tipo_bloque", "5col")
    ancho_bloque = bloques_dia[0].get("ancho_bloque", 6)

    return {
        "fila_fechas": fila_fechas,
        "fila_subtitulos": fila_subtitulos,
        "fila_inicio_datos": fila_inicio_datos,
        "col_nombre": col_nombre,
        "col_legajo": col_legajo,
        "bloques_dia": bloques_dia,
        "columnas_resumen": columnas_resumen,
        "tipo_bloque": tipo_bloque,
        "ancho_bloque": ancho_bloque,
    }

def detectar_hoja_planilla(wb):
    mejores = []
    for ws in wb.worksheets:
        try:
            estructura = detectar_estructura_planilla(ws)
            score = len(estructura["bloques_dia"])
            mejores.append((score, ws.title, estructura))
        except Exception:
            continue

    if not mejores:
        raise ValueError("No encontré ninguna hoja con estructura de planilla compatible.")

    mejores.sort(reverse=True, key=lambda x: x[0])
    _, hoja, estructura = mejores[0]
    return wb[hoja], estructura


# =========================================================
# FECHAS / FERIADOS
# =========================================================
DIAS_ES = {
    0: "lunes",
    1: "martes",
    2: "miércoles",
    3: "jueves",
    4: "viernes",
    5: "sábado",
    6: "domingo",
}
MESES_ES = {
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}

def formatear_fecha_es(fecha):
    return f"{DIAS_ES[fecha.weekday()]}, {fecha.day} de {MESES_ES[fecha.month]} de {fecha.year}"

def ultimo_dia_mes(anio, mes):
    return calendar.monthrange(anio, mes)[1]

def obtener_feriados(anio, mes):
    feriados = {}
    try:
        kwargs = {"years": [anio]}
        if SUBDIV_FERIADOS:
            kwargs["subdiv"] = SUBDIV_FERIADOS
        cal = holidays.country_holidays(PAIS_FERIADOS, **kwargs)
        for f, nombre in cal.items():
            if f.year == anio and f.month == mes:
                feriados[f] = nombre
    except Exception as e:
        print(f"Aviso: no pude cargar feriados automáticos ({e}). Se usarán solo los manuales.")

    for f in FERIADOS_ADICIONALES:
        try:
            d = date.fromisoformat(f)
            if d.year == anio and d.month == mes:
                feriados[d] = "FERIADO MANUAL"
        except Exception:
            pass

    for f in FERIADOS_EXCLUIDOS:
        try:
            d = date.fromisoformat(f)
            feriados.pop(d, None)
        except Exception:
            pass

    return feriados


# =========================================================
# EXCEL / FÓRMULAS
# =========================================================
def aplicar_tolerancia(ts):
    if pd.isna(ts):
        return ts

    ts = pd.Timestamp(ts)

    if ts.minute <= TOLERANCIA_MINUTOS:
        return ts.replace(minute=0, second=0, microsecond=0)

    return ts.replace(second=0, microsecond=0)

def formula_suma(refs):
    if not refs:
        return "=0"
    return "=SUM(" + ",".join(refs) + ")"

def construir_listas_resumen(estructura, anio, mes, feriados):
    dias_mes = ultimo_dia_mes(anio, mes)

    refs = {
        "q1": {"normales": [], "50": [], "100": []},
        "q2": {"normales": [], "50": [], "100": []},
    }

    for dia in range(1, min(dias_mes, len(estructura["bloques_dia"])) + 1):
        bloque = estructura["bloques_dia"][dia - 1]
        fecha = date(anio, mes, dia)
        q = "q1" if dia <= 15 else "q2"

        ref_horas = f"{get_column_letter(bloque['horas'])}{{fila}}"
        ref_ex = f"{get_column_letter(bloque['horas_ex'])}{{fila}}"

        es_feriado = fecha in feriados
        wd = fecha.weekday()  # 0=lun ... 6=dom

        if es_feriado or wd == 6:
            refs[q]["100"].append(ref_horas)
            refs[q]["100"].append(ref_ex)
        elif wd == 5:  # sábado
            refs[q]["normales"].append(ref_horas)
            refs[q]["100"].append(ref_ex)
        else:  # lunes a viernes
            refs[q]["normales"].append(ref_horas)
            refs[q]["50"].append(ref_ex)

    return refs


def reconfigurar_mes_en_planilla(ws, estructura, anio, mes, feriados):
    fila_fechas = estructura["fila_fechas"]
    dias_mes = ultimo_dia_mes(anio, mes)

    for i, bloque in enumerate(estructura["bloques_dia"], start=1):
        visible = i <= dias_mes
        fecha = date(anio, mes, i) if visible else None

        ws.cell(fila_fechas, bloque["base_col"]).value = formatear_fecha_es(fecha) if fecha else None

        ancho = bloque.get("ancho_bloque", estructura.get("ancho_bloque", 6))
        for c in range(bloque["base_col"], bloque["base_col"] + ancho):
            ws.column_dimensions[get_column_letter(c)].hidden = not visible

def limpiar_celdas_carga(ws, estructura, fila):
    for bloque in estructura["bloques_dia"]:
        for clave in ("entrada", "salida", "ausente"):
            celda = ws.cell(fila, bloque[clave])
            celda.value = None
            celda.fill = NO_FILL

def forzar_recalculo(wb):
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

def actualizar_formulas_resumen(ws, estructura, anio, mes, feriados):
    refs = construir_listas_resumen(estructura, anio, mes, feriados)
    res = estructura["columnas_resumen"]
    fila_inicio = estructura["fila_inicio_datos"]
    col_nombre = estructura["col_nombre"]

    for fila in range(fila_inicio, ws.max_row + 1):
        nombre = ws.cell(fila, col_nombre).value
        if nombre is None or str(nombre).strip() == "":
            continue

        q1_norm = [r.format(fila=fila) for r in refs["q1"]["normales"]]
        q1_50 = [r.format(fila=fila) for r in refs["q1"]["50"]]
        q1_100 = [r.format(fila=fila) for r in refs["q1"]["100"]]

        q2_norm = [r.format(fila=fila) for r in refs["q2"]["normales"]]
        q2_50 = [r.format(fila=fila) for r in refs["q2"]["50"]]
        q2_100 = [r.format(fila=fila) for r in refs["q2"]["100"]]

        ws.cell(fila, res["q1"]["normales"]).value = formula_suma(q1_norm)
        ws.cell(fila, res["q1"]["50"]).value = formula_suma(q1_50)
        ws.cell(fila, res["q1"]["100"]).value = formula_suma(q1_100)

        ws.cell(fila, res["q2"]["normales"]).value = formula_suma(q2_norm)
        ws.cell(fila, res["q2"]["50"]).value = formula_suma(q2_50)
        ws.cell(fila, res["q2"]["100"]).value = formula_suma(q2_100)

        c_q1_n = f"{get_column_letter(res['q1']['normales'])}{fila}"
        c_q1_50 = f"{get_column_letter(res['q1']['50'])}{fila}"
        c_q1_100 = f"{get_column_letter(res['q1']['100'])}{fila}"

        c_q2_n = f"{get_column_letter(res['q2']['normales'])}{fila}"
        c_q2_50 = f"{get_column_letter(res['q2']['50'])}{fila}"
        c_q2_100 = f"{get_column_letter(res['q2']['100'])}{fila}"

        ws.cell(fila, res["totales"]["normales"]).value = f"={c_q1_n}+{c_q2_n}"
        ws.cell(fila, res["totales"]["50"]).value = f"={c_q1_50}+{c_q2_50}"
        ws.cell(fila, res["totales"]["100"]).value = f"={c_q1_100}+{c_q2_100}"



def _f_mon_jue_horas(fila, col_entrada, col_salida):
    e = get_column_letter(col_entrada)
    s = get_column_letter(col_salida)
    return (
        f'=IF(AND({e}{fila}<>"",{s}{fila}<>""),'
        f'MIN(9,IF(({s}{fila}-IF({e}{fila}<8/24,8/24,{e}{fila}))*24>=8.5,9,'
        f'HOUR({s}{fila}-IF({e}{fila}<8/24,8/24,{e}{fila}))+'
        f'IF(MINUTE({s}{fila}-IF({e}{fila}<8/24,8/24,{e}{fila}))>30,1,0))),"")'
    )

def _f_mon_jue_50(fila, col_entrada, col_salida, col_ausente):
    e = get_column_letter(col_entrada)
    s = get_column_letter(col_salida)
    a = get_column_letter(col_ausente)
    return (
        f'=IF({e}{fila}="","",IF({s}{fila}="","",IF({a}{fila}="AUSENTE","",'
        f'IF(MOD({e}{fila},1)>=23/24,1,0)+'
        f'IF(MOD({s}{fila},1)<17.45/24+1/1440,0,'
        f'IF(MOD({s}{fila},1)<18.45/24+1/1440,1,'
        f'IF(MOD({s}{fila},1)<19.45/24+1/1440,2,'
        f'IF(MOD({s}{fila},1)<20.45/24+1/1440,3,'
        f'IF(MOD({s}{fila},1)<21.45/24+1/1440,4,5))))))))'
    )

def _f_viernes_horas(fila, col_entrada, col_salida):
    e = get_column_letter(col_entrada)
    s = get_column_letter(col_salida)
    return (
        f'=IF(AND({e}{fila}<>"",{s}{fila}<>""),'
        f'MIN(8,IF(({s}{fila}-{e}{fila})*24>=7.5,8,'
        f'HOUR({s}{fila}-{e}{fila})+IF(MINUTE({s}{fila}-{e}{fila})>30,1,0))),"")'
    )

def _f_viernes_50(fila, col_entrada, col_salida, col_ausente):
    e = get_column_letter(col_entrada)
    s = get_column_letter(col_salida)
    a = get_column_letter(col_ausente)
    return (
        f'=IF({e}{fila}="","",IF({s}{fila}="","",IF({a}{fila}="AUSENTE","",'
        f'IF(MOD({e}{fila},1)<=7/24,1,0)+'
        f'IF(MOD({s}{fila},1)<16/24+1/1440,0,'
        f'IF(MOD({s}{fila},1)<17/24+1/1440,1,'
        f'IF(MOD({s}{fila},1)<18/24+1/1440,2,'
        f'IF(MOD({s}{fila},1)<19/24+1/1440,3,'
        f'IF(MOD({s}{fila},1)<20/24+1/1440,4,5))))))))'
    )

def _f_sabado_50(fila, col_entrada, col_salida):
    e = get_column_letter(col_entrada)
    s = get_column_letter(col_salida)
    return (
        f'=IF(AND({e}{fila}<>"",{s}{fila}<>""),'
        f'MIN(4,IF(({s}{fila}-{e}{fila})*24>=3.5,4,'
        f'HOUR({s}{fila}-{e}{fila})+IF(MINUTE({s}{fila}-{e}{fila})>30,1,0))),"")'
    )

def _f_sabado_100(fila, col_salida):
    s = get_column_letter(col_salida)
    return f'=IF({s}{fila}*24<13,0,ROUND({s}{fila}*24,0)-13)'

def _f_domingo_100(fila, col_entrada, col_salida):
    e = get_column_letter(col_entrada)
    s = get_column_letter(col_salida)
    return (
        f'=IF(AND({e}{fila}<>"",{s}{fila}<>""),'
        f'MIN(IF(({s}{fila}-{e}{fila})*24>=7.5,'
        f'HOUR({s}{fila}-{e}{fila})+IF(MINUTE({s}{fila}-{e}{fila})>30,1,0))),"")'
    )

def actualizar_formulas_diarias(ws, estructura, anio, mes, feriados):
    if estructura.get("tipo_bloque") != "6col":
        return

    dias_mes = ultimo_dia_mes(anio, mes)
    fila_inicio = estructura["fila_inicio_datos"]
    col_nombre = estructura["col_nombre"]

    for fila in range(fila_inicio, ws.max_row + 1):
        nombre = ws.cell(fila, col_nombre).value
        if nombre is None or str(nombre).strip() == "":
            continue

        for dia in range(1, min(dias_mes, len(estructura["bloques_dia"])) + 1):
            bloque = estructura["bloques_dia"][dia - 1]
            fecha = date(anio, mes, dia)
            wd = fecha.weekday()
            es_feriado = fecha in feriados

            c_h = ws.cell(fila, bloque["horas"])
            c_50 = ws.cell(fila, bloque["horas_50"])
            c_100 = ws.cell(fila, bloque["horas_100"])

            if es_feriado or wd == 6:
                c_h.value = None
                c_50.value = None
                c_100.value = _f_domingo_100(fila, bloque["entrada"], bloque["salida"])
            elif wd in {0, 1, 2, 3}:
                c_h.value = _f_mon_jue_horas(fila, bloque["entrada"], bloque["salida"])
                c_50.value = _f_mon_jue_50(fila, bloque["entrada"], bloque["salida"], bloque["ausente"])
                c_100.value = None
            elif wd == 4:
                c_h.value = _f_viernes_horas(fila, bloque["entrada"], bloque["salida"])
                c_50.value = _f_viernes_50(fila, bloque["entrada"], bloque["salida"], bloque["ausente"])
                c_100.value = None
            elif wd == 5:
                c_h.value = None
                c_50.value = _f_sabado_50(fila, bloque["entrada"], bloque["salida"])
                c_100.value = _f_sabado_100(fila, bloque["salida"])


# =========================================================
# HOJAS AUXILIARES
# =========================================================
def borrar_hojas_auxiliares(wb):
    for hoja in ["COINCIDENCIAS_NOMBRES", "RESUMEN", "INCIDENCIAS", "REVISION_MANUAL", "PARAMETROS"]:
        if hoja in wb.sheetnames:
            wb.remove(wb[hoja])

def crear_hojas_auxiliares(wb, coincidencias_rows, resumen_rows, incidencias_rows, revision_rows, parametros_rows):
    ws_match = wb.create_sheet("COINCIDENCIAS_NOMBRES")
    ws_match.append([
        "NOMBRE_PLANILLA",
        "LEG_PLANILLA",
        "NOMBRE_FICHAJES",
        "LEG_FICHAJES",
        "METODO_MATCH",
        "SCORE",
        "SCORE_ORIGINAL",
        "ESTADO",
        "MATCH_MANUAL",
        "SIM_BASE",
        "SIM_TOKEN_SORT",
        "SIM_TOKENS",
        "SIM_APELLIDO_PRIMERNOMBRE",
        "CANDIDATOS_POR_LEGAJO",
    ])
    for row in coincidencias_rows:
        ws_match.append(row)

    ws_resumen = wb.create_sheet("RESUMEN")
    ws_resumen.append([
        "PERSONA",
        "LEG_PLANILLA",
        "COINCIDENCIA_FICHAJES",
        "LEG_FICHAJES",
        "METODO_MATCH",
        "SCORE",
        "SCORE_ORIGINAL",
        "AUSENCIAS",
        "TARDANZAS",
    ])
    for row in resumen_rows:
        ws_resumen.append(row)

    ws_incid = wb.create_sheet("INCIDENCIAS")
    ws_incid.append(["PERSONA", "LEG_PLANILLA", "DIA", "PROBLEMA", "SCORE_MATCH", "METODO_MATCH"])
    for row in incidencias_rows:
        ws_incid.append(row)

    ws_rev = wb.create_sheet("REVISION_MANUAL")
    ws_rev.append(["REFERENCIA", "LEG_PLANILLA", "COINCIDENCIA / DIA", "SCORE", "MOTIVO"])
    for row in revision_rows:
        ws_rev.append(row)

    ws_param = wb.create_sheet("PARAMETROS")
    ws_param.append(["CLAVE", "VALOR"])
    for row in parametros_rows:
        ws_param.append(row)

    for hoja in [ws_match, ws_resumen, ws_incid, ws_rev, ws_param]:
        for cell in hoja[1]:
            cell.font = Font(bold=True)


# =========================================================
# PROCESO PRINCIPAL
# =========================================================
def procesar_periodo(archivo_planilla, agrupado_mes, detalle_mes, periodo, descargar_en_colab=True):
    anio = int(periodo.year)
    mes = int(periodo.month)

    wb = load_workbook(archivo_planilla, data_only=False)
    ws, estructura = detectar_hoja_planilla(wb)

    forzar_recalculo(wb)
    borrar_hojas_auxiliares(wb)

    feriados = obtener_feriados(anio, mes)
    reconfigurar_mes_en_planilla(ws, estructura, anio, mes, feriados)
    if estructura.get("tipo_bloque") == "6col":
        actualizar_formulas_diarias(ws, estructura, anio, mes, feriados)
    else:
        actualizar_formulas_resumen(ws, estructura, anio, mes, feriados)

    fila_inicio = estructura["fila_inicio_datos"]
    col_nombre = estructura["col_nombre"]
    col_legajo = estructura.get("col_legajo")
    dias_mes = ultimo_dia_mes(anio, mes)
    bloques_validos = estructura["bloques_dia"][:dias_mes]

    nombres_fichajes = sorted(agrupado_mes["NombreLimpio"].dropna().unique().tolist())

    detalle_personas = (
        detalle_mes[["NombreLimpio", "NombreOriginal", "LegajoFichaje"]]
        .drop_duplicates()
        .copy()
    )
    detalle_personas["LegajoFichaje"] = detalle_personas["LegajoFichaje"].apply(normalizar_legajo)

    legajo_a_candidatos = {}
    for leg, sub in detalle_personas.dropna(subset=["LegajoFichaje"]).groupby("LegajoFichaje"):
        legajo_a_candidatos[leg] = sorted(sub["NombreLimpio"].dropna().unique().tolist())

    coincidencias_rows = []
    resumen_rows = []
    incidencias_rows = []
    revision_rows = []

    for fila in range(fila_inicio, ws.max_row + 1):
        nombre_planilla = ws.cell(fila, col_nombre).value

        if nombre_planilla is None or str(nombre_planilla).strip() == "":
            continue

        limpiar_celdas_carga(ws, estructura, fila)

        nombre_planilla_str = str(nombre_planilla).strip()
        legajo_planilla = normalizar_legajo(ws.cell(fila, col_legajo).value) if col_legajo else None

        resultado_match = resolver_coincidencia_persona(
            nombre_planilla=nombre_planilla_str,
            legajo_planilla=legajo_planilla,
            candidatos_fichajes=nombres_fichajes,
            legajo_a_candidatos=legajo_a_candidatos,
        )

        nombre_match = resultado_match["nombre_match"]
        score = resultado_match["score"]
        score_original = resultado_match.get("score_original", score)
        detalle_match = resultado_match.get("detalle", {}) or {}
        metodo_match = detalle_match.get("metodo", "nombre")
        leg_fichaje = detalle_match.get("legajo_fichajes")

        if nombre_match:
            nombre_fichaje_mostrar = agrupado_mes.loc[
                agrupado_mes["NombreLimpio"] == nombre_match, "NombreFichaje"
            ].iloc[0]
        else:
            nombre_fichaje_mostrar = ""

        estado = "OK" if resultado_match["aceptado"] else "REVISAR"

        coincidencias_rows.append([
            nombre_planilla_str,
            legajo_planilla,
            nombre_fichaje_mostrar,
            leg_fichaje,
            metodo_match,
            round(score, 4),
            round(score_original, 4),
            estado,
            detalle_match.get("manual", False),
            detalle_match.get("base", ""),
            detalle_match.get("token_sort", ""),
            detalle_match.get("tokens", ""),
            detalle_match.get("apellido_primernombre", ""),
            detalle_match.get("candidatos_legajo", ""),
        ])

        if not resultado_match["aceptado"]:
            ws.cell(fila, col_nombre).fill = FILL_NO_MATCH
            revision_rows.append([
                nombre_planilla_str,
                legajo_planilla,
                nombre_fichaje_mostrar,
                round(score_original, 4),
                "SIN COINCIDENCIA CONFIABLE",
            ])
            incidencias_rows.append([
                nombre_planilla_str,
                legajo_planilla,
                "-",
                "SIN COINCIDENCIA CONFIABLE (NOMBRE/LEGAJO)",
                round(score_original, 4),
                metodo_match,
            ])
            continue
        elif resultado_match["dudoso"]:
            ws.cell(fila, col_nombre).fill = FILL_DUDOSO
            revision_rows.append([
                nombre_planilla_str,
                legajo_planilla,
                nombre_fichaje_mostrar,
                round(score_original, 4),
                "COINCIDENCIA DUDOSA",
            ])
        else:
            ws.cell(fila, col_nombre).fill = NO_FILL

        ausencias = 0
        tardanzas = 0

        for dia, bloque in enumerate(bloques_validos, start=1):
            fecha_actual = date(anio, mes, dia)
            es_feriado = fecha_actual in feriados
            wd = fecha_actual.weekday()

            reg = agrupado_mes[
                (agrupado_mes["NombreLimpio"] == nombre_match) &
                (agrupado_mes["Fecha"] == fecha_actual)
            ]

            col_entrada = bloque["entrada"]
            col_salida = bloque["salida"]
            col_ausente = bloque["ausente"]

            if reg.empty:
                if wd in DIAS_LABORABLES_AUSENCIA and not es_feriado:
                    ws.cell(fila, col_ausente).value = "AUSENTE"
                    ws.cell(fila, col_ausente).fill = FILL_INCIDENCIA
                    ausencias += 1
                    incidencias_rows.append([nombre_planilla_str, legajo_planilla, dia, "AUSENTE", round(score_original, 4), metodo_match])
                continue

            entrada = reg.iloc[0]["Entrada"]
            salida = reg.iloc[0]["Salida"]
            cantidad = int(reg.iloc[0]["CantidadFichajes"])

            entrada_ajustada = aplicar_tolerancia(entrada)

            c_entrada = ws.cell(fila, col_entrada)
            c_entrada.value = entrada_ajustada.to_pydatetime().time()
            c_entrada.number_format = "hh:mm"
            c_entrada.fill = FILL_CARGADO

            if entrada == salida:
                c_salida = ws.cell(fila, col_salida)
                c_salida.value = None
                incidencias_rows.append([nombre_planilla_str, dia, "SOLO UN FICHAJE", round(score, 4)])
                revision_rows.append([nombre_planilla_str, dia, round(score, 4), "SOLO UN FICHAJE"])
            else:
                c_salida = ws.cell(fila, col_salida)
                c_salida.value = pd.Timestamp(salida).to_pydatetime().time()
                c_salida.number_format = "hh:mm"
                c_salida.fill = FILL_CARGADO

            ws.cell(fila, col_ausente).value = None
            ws.cell(fila, col_ausente).fill = NO_FILL

            if entrada_ajustada.hour > 8 or (
                entrada_ajustada.hour == 8 and entrada_ajustada.minute > TOLERANCIA_MINUTOS
            ):
                tardanzas += 1
                incidencias_rows.append([
                    nombre_planilla_str,
                    legajo_planilla,
                    dia,
                    f"LLEGADA TARDE ({entrada_ajustada.strftime('%H:%M')})",
                    round(score_original, 4),
                    metodo_match,
                ])

            if cantidad == 1:
                incidencias_rows.append([
                    nombre_planilla_str,
                    legajo_planilla,
                    dia,
                    "SOLO UN FICHAJE",
                    round(score_original, 4),
                    metodo_match,
                ])

        resumen_rows.append([
            nombre_planilla_str,
            legajo_planilla,
            nombre_fichaje_mostrar,
            leg_fichaje,
            metodo_match,
            round(score, 4),
            round(score_original, 4),
            ausencias,
            tardanzas,
        ])

    parametros_rows = [
        ["ARCHIVO_PLANILLA_BASE", os.path.basename(archivo_planilla)],
        ["PERIODO", f"{anio:04d}-{mes:02d}"],
        ["DIAS_DEL_MES", dias_mes],
        ["FERIADOS_DETECTADOS", ", ".join(
            [f"{d.isoformat()} - {feriados[d]}" for d in sorted(feriados)]
        ) if feriados else "NINGUNO"],
        ["UMBRAL_MATCH", UMBRAL_MATCH],
        ["UMBRAL_DUDOSO", UMBRAL_DUDOSO],
        ["UMBRAL_MINIMO_SIM_POR_LEGAJO", UMBRAL_MINIMO_SIM_POR_LEGAJO],
        ["TOLERANCIA_MINUTOS", TOLERANCIA_MINUTOS],
        ["PROCESAR_TODOS_LOS_MESES", PROCESAR_TODOS_LOS_MESES],
    ]

    crear_hojas_auxiliares(
        wb=wb,
        coincidencias_rows=coincidencias_rows,
        resumen_rows=resumen_rows,
        incidencias_rows=incidencias_rows,
        revision_rows=revision_rows,
        parametros_rows=parametros_rows,
    )

    base_planilla = Path(archivo_planilla).stem
    nombre_salida = f"{PREFIJO_SALIDA}_{base_planilla}_{anio:04d}_{mes:02d}.xlsx"
    wb.save(nombre_salida)

    print(f"Archivo generado: {nombre_salida}")
    print("Al abrirlo en Excel, las fórmulas quedan listas para recalcularse automáticamente.")

    if descargar_en_colab and files is not None:
        try:
            files.download(nombre_salida)
        except Exception:
            pass

    return nombre_salida

def procesar_archivos(archivo_planilla, archivo_fichajes, descargar_en_colab=True):
    agrupado, detalle = leer_fichajes(archivo_fichajes)

    periodos = sorted(agrupado["Periodo"].dropna().unique())
    if not periodos:
        raise ValueError("No hay períodos válidos en el archivo de fichajes.")

    if not PROCESAR_TODOS_LOS_MESES:
        periodos = [max(periodos)]

    salidas = []
    for periodo in periodos:
        agrupado_mes = agrupado[agrupado["Periodo"] == periodo].copy()
        detalle_mes = detalle[detalle["Periodo"] == periodo].copy()
        salida = procesar_periodo(
            archivo_planilla=archivo_planilla,
            agrupado_mes=agrupado_mes,
            detalle_mes=detalle_mes,
            periodo=periodo,
            descargar_en_colab=descargar_en_colab,
        )
        salidas.append(salida)

    return salidas

def ejecutar_desde_colab():
    if files is None:
        raise RuntimeError("Este modo requiere Google Colab.")
    print("Subí los Excel de la obra:")
    print("- uno con fichajes")
    print("- uno con la planilla base")
    uploaded = files.upload()
    archivos = list(uploaded.keys())

    archivo_planilla, archivo_fichajes, diagnostico = detectar_archivos(archivos)

    print("\nArchivos detectados:")
    print(f"- Planilla base : {archivo_planilla}")
    print(f"- Fichajes      : {archivo_fichajes}")

    print("\nDiagnóstico de detección:")
    for item in diagnostico:
        print(
            f"  {item['path']} | score_planilla={item['score_planilla']} | "
            f"score_fichajes={item['score_fichajes']}"
        )

    salidas = procesar_archivos(
        archivo_planilla=archivo_planilla,
        archivo_fichajes=archivo_fichajes,
        descargar_en_colab=True,
    )

    print("\nProceso finalizado.")
    for s in salidas:
        print("-", s)

# =========================================================
# EJECUCIÓN
# =========================================================
if __name__ == "__main__":
    if files is None:
        print("Modo local detectado. Importá este archivo y llamá a:")
        print("procesar_archivos('planilla.xlsx', 'fichajes.xls', descargar_en_colab=False)")
    else:
        ejecutar_desde_colab()
